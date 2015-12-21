#!/usr/bin/python3
import sys
import os
import openpyxl
from collections import OrderedDict

__author__ = 'Nika Tsanakshvili'
__email__ = "nikatsanka@gmail.com"
__version__ = "1.3"
__copyright__ = "Copyright (C) 2015 Nika Tsankashvili"
__license__ = "Public Domain"

# Search Range for keys
START_COL = 1
END_COL = 7


def search_xl():
    # open the workbook
    # Before assignments
    global start_row, wb, values_full
    column_sums = OrderedDict()
    sheet_sum = list()
    sheet_sum_of_col = list()

    # if length of the argv is more than 1 then the argument is passed.
    if len(sys.argv) > 1:
        # path should be at 1 because .py is at 0
        workbook = sys.argv[1]
    else:
        workbook = input("Please enter name of the workbook: ")

    input_val = True
    while input_val:
        try:
            wb = openpyxl.load_workbook(workbook, read_only=True)
            input_val = False
        except FileNotFoundError:
            print("Specified workbook cannot be located.")
            workbook = input("Please enter name of the workbook: ")

    # get the search value.
    search_val = input("Search bacteria: ")

    print("Available Sheets")
    # print sheets
    sheets = wb.get_sheet_names()
    print(sheets)

    # find home directory / desktop.
    desktop_dir = os.path.join(os.path.expanduser('~'), 'Desktop')

    # create an output file
    output_file = os.path.join(desktop_dir, search_val + "-Results-From-" + workbook + ".txt")

    with open(output_file, 'w') as output:
        # keep checking sheets
        user_ = True
        while user_:
            # get working sheet.
            sheet_by_name = input("Which sheet would you like to search: ")
            # if chosen sheet is in the workbook then do the following
            if sheet_by_name in sheets:
                # set the sheet
                sheet = wb.get_sheet_by_name(sheet_by_name)

                # get maximum number of rows in a sheet.
                max_row = sheet.max_row
                max_col = sheet.max_column - 1

                # find the starting point
                # start looking for # sign in the first column
                for k in sheet.columns[0]:
                    # start checking
                    cell_val = k.value
                    if cell_val == '#':
                        start_row = k.row
                        break
                # save number of results
                count_res = 0

                table_header = ''
                sum_header = ''
                values = ''

                output.write("\n" + sheet_by_name + "\n")

                # print table headers
                # 1. reserve enough spaces for search value
                search_len = len(search_val) + 12
                for col in range(END_COL, max_col):
                    table_header += '{:<{}s}'.format(sheet.cell(row=start_row, column=col).value,
                                                     len(sheet.cell(row=start_row, column=col).value) + 4)
                    # store locations as keys in the dictionary.
                    column_sums[str(sheet.cell(row=start_row, column=col).value)] = []

                table_header_full = '{:<{}s}'.format(' ', search_len) + table_header + '\n'

                output.write(table_header_full)

                # start looking for search value
                # 1 to 7 columns that's first "#" sign
                for i in range(START_COL, END_COL):
                    # start_row to max_row. STEP 1
                    for j in range(start_row, max_row, 1):
                        # check for search value
                        if search_val == sheet.cell(row=j, column=i).value:
                            # reset values
                            values = ''
                            values_full = ''

                            # print the entire row.
                            # create a string first
                            for in_col in range(END_COL, max_col):
                                # get spaces from columns
                                col_len = len(sheet.cell(row=start_row, column=in_col).value)
                                values += '{:<{}.12}'.format(str(sheet.cell(row=j, column=in_col).value), col_len + 4)

                                # store each value in the dictionary according to the key/name/location.
                                col_val = column_sums[str(sheet.cell(row=start_row, column=in_col).value)]
                                # append at the end of the list
                                col_val.append(sheet.cell(row=j, column=in_col).value)
                                column_sums[str(sheet.cell(row=start_row, column=in_col).value)] = col_val

                            values_full = '{:<{}s}'.format(search_val + " " + str(count_res + 1) + ": " +
                                                           sheet.cell(row=j, column=i).coordinate, search_len) + values
                            output.write(values_full + '\n')

                            count_res += 1
                # clear the list
                del sheet_sum_of_col[:]

                # sum up columns
                for location in column_sums:
                    # get the column value
                    column_values = column_sums[location]
                    # sum up the column
                    sum_of_columns = sum(column_values)
                    # append to overall sum
                    sheet_sum.append(sum_of_columns)
                    # append to sheet sum
                    sheet_sum_of_col.append(sum_of_columns)
                    # put the value back for printing.
                    column_sums[location] = sum_of_columns

                # prints _s
                for n in range(0, len(values_full)):
                    output.write("_")

                counter = 0
                # and values
                for col in range(END_COL, max_col):
                    sum_header += '{:<{}f}'.format(sheet_sum_of_col[counter],
                                                     len(sheet.cell(row=start_row, column=col).value) + 4)
                    counter += 1

                sum_header_names = '{:<{}s}'.format("Sum: ", search_len) + table_header
                sum_header_full = '{:<{}s}'.format(' ', search_len) + sum_header
                sum_total = '{:<{}s}'.format("Sheet Overall: ", search_len) + str(sum(sheet_sum_of_col))

                output.write("\n" + sum_header_names)
                output.write("\n" + sum_header_full)
                output.write("\n" + sum_total + "\n")
                print("Search Completed")
                # ask the user for more sheets
                ans = input("Would you like to search another sheet? (Y/n): ")
                if ans == 'N' or ans == 'n':
                    output.write("\n" + search_val + " Taxa Abundance: " + str(sum(sheet_sum)))
                    print("Exiting...")
                    user_ = False
                else:
                    print("Available Sheets")
                    # print sheets
                    sheets = wb.get_sheet_names()
                    print(sheets)
                    # else it will ask the user to enter the sheet name again.
            else:
                print("\'", sheet_by_name, "\' is not in the workbook.", sep='')


if __name__ == "__main__":
    if sys.version_info > (3, 4):
        search_xl()
    else:
        print("This software requires Python 3.4 or higher to run." +
              "\nYou can download the latest version of Python from: https://www.python.org/")
