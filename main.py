"""
Open and read the cells of an Excel document with the openpyxl module.
Calculate all the tract and population data and store it in a data structure.
Write the data structure to a text file with the .py extension using the pprint module.
"""

import sys
import openpyxl
from openpyxl.utils import get_column_letter


def get_files():
    # open the workbook
    # Before assignments
    global start_row, wb

    # if length of the argv is more than 1 then the argument is passed.
    if len(sys.argv) > 1:
        # path should be at 1 because .py is at 0
        workbook = sys.argv[1]
    else:
        workbook = input("Please enter name of the workbook: ")

    input_val = True
    while input_val:
        try:
            wb = openpyxl.load_workbook(workbook, read_only=True)
            input_val = False
        except FileNotFoundError:
            print("Specified workbook cannot be located.")
            workbook = input("Please enter name of the workbook: ")

    print("Available Sheets")
    # print sheets
    sheets = wb.get_sheet_names()
    print(sheets)
    # get the search value.
    search_val = input("What are you looking for? ")
    # ask usr for more sheets.
    user_ = True
    # keep checking sheets
    while user_:
        # get working sheet.
        sheet_by_name = input("Which sheet would you like to work with: ")
        # if chosen sheet is in the workbook then do the following
        if sheet_by_name in sheets:
            # set the sheet
            sheet = wb.get_sheet_by_name(sheet_by_name)

            # get maximum number of rows in a sheet.
            max_row = sheet.max_row

            # find the starting point
            # start looking for # sign in the first column
            for k in sheet.columns[0]:
                # start checking
                cell_val = k.value
                if cell_val == '#':
                    start_row = k.row
                    break
            # save number of results
            count_res = 0

            # start looking for search value
            # 1 to 7 columns that's first 1
            for i in range(1, 7):
                # start_row to max_row. STEP 1
                for j in range(start_row, max_row, 1):
                    # check for search value
                    if search_val == sheet.cell(row=j, column=i).value:
                        print("[ Found at row:", j, "column:", get_column_letter(i), "]")
                        count_res += 1

            print("Total number of \'", search_val, "\' found: ", count_res, sep='')
            ans = input("Would you like to search another sheet? (Y/n): ")
            if ans == 'N' or ans == 'n':
                user_ = False
            else:
                print("Available Sheets")
                # print sheets
                sheets = wb.get_sheet_names()
                print(sheets)
            # else it will ask the user to enter the sheet name again.
        else:
            print("\'", sheet_by_name, "\' is not in the workbook.", sep='')

if __name__ == "__main__":
    get_files()
